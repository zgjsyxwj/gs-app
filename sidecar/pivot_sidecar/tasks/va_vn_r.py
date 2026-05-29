"""瓦里安越南 Payroll 报告加工 (VA-VN-PAYROLL-REPORT)

按 station (GL / 13th / Variance) 分派处理。前端每个 station 单独发起一次
run；options.station 决定走哪条分支。GL / 13th / Variance 均已实现真实逻辑。

供应商发来的 xlsx 用 ``vnpayroll`` 密码加密；解密后用 openpyxl 原地改若干
单元格（保留样式），再以同样的密码重新加密写回。

GL / 13th 规则一致，只是列名不同：
  - GL 报告：列名 "G/L Acc"
  - 13th 报告：列名 "G/L Code"（也接受 "GL Code"）

对每一行：当 GL 编码以 "2" 开头 → BusinessArea=2000、ProfitCenter=10000000、
CostCenter 清空。以 "6" 开头则不动（供应商已填好 CostCenter）。

Variance：前端传入两个文件——Payroll 报告（input_path）与 Variance 报告
(options["input2"])。按 GID 关联，把 Payroll 的 Start date / Last working date
两列插入到 Variance 的 Full Name 之后（第 5、6 列），再加密写回
VarianVN_Variance_<period>_CDP.xlsx。
"""
from __future__ import annotations
import io
import os
import time
from copy import copy
from pathlib import Path
from typing import Iterator

import msoffcrypto
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from .base import TaskBase, TaskEvent
from ..ipc import LogEvent, ProgressEvent
from ..xlsx import ExcelError, open_xlsx


PASSWORD = "vnpayroll"

# 列名候选（不同年份的模板可能存在轻微差异）
GL_CODE_HEADERS = ("G/L Acc", "G/L Code", "GL Code", "GL Acc")
BA_HEADER = "BusinessArea"
CC_HEADER = "CostCenter"
PC_HEADER = "ProfitCenter"

BA_VALUE = 2000
PC_VALUE = 10000000

# Variance station：按 GID 把 Payroll 这两列插入到 Variance 的 Full Name 之后。
# Payroll 实际表头形如 "Start date (dd/mm/yyyy)"，用「包含」匹配以兼容年份差异。
PR_START_HEADER = "Start date"
PR_LAST_HEADER = "Last working date"
OUT_START_HEADER = "Start date (dd/mm/yyyy)"
OUT_LAST_HEADER = "Last working date (dd/mm/yyyy)"

# 无 station 的旧调用仍走 placeholder（写随机字节），保留前端 UI 可用性
VARIANCE_PLACEHOLDER_BYTES = 142 * 1024


def _save_encrypted(wb: Workbook, dst: Path, password: str) -> None:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with open(dst, "wb") as fout:
        of = msoffcrypto.OfficeFile(buf)
        of.load_key(password=password)
        of.encrypt(password, fout)


def _find_col(ws, candidates: tuple[str, ...]) -> int | None:
    headers = {(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
    for name in candidates:
        if name in headers:
            return headers[name]
    return None


def _starts_with(value, prefix: str) -> bool:
    if value is None:
        return False
    s = str(value).strip()
    return s.startswith(prefix)


def _find_header_col(ws, candidates: tuple[str, ...], *, contains: bool = False,
                     max_scan_rows: int = 20) -> int | None:
    """在前 max_scan_rows 行里按阅读顺序找首个命中表头的列（1-based）。

    GL/13th 表头在第 1 行，故 ``_find_col`` 够用；Variance 用到的 Payroll 表头
    跨第 9/11 行，这里扫描多行以兼容。
    """
    upper = min(ws.max_row, max_scan_rows)
    for r in range(1, upper + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip()
            for cand in candidates:
                if (cand.lower() in s.lower()) if contains else (s == cand):
                    return c
    return None


def _find_payroll_sheet(wb):
    """含 GID 且含 Start date 的工作表即 Payroll 主表（避开 Payment Notice）。"""
    for ws in wb.worksheets:
        if (_find_header_col(ws, ("GID",)) is not None
                and _find_header_col(ws, (PR_START_HEADER,), contains=True) is not None):
            return ws
    return None


def _build_payroll_map(ws) -> dict[str, tuple]:
    """GID → (start_value, start_fmt, last_value, last_fmt)。

    供应商表里 Start date 是文本（如 '01-07-2019'），Last working date 是数字 0
    且 number_format 为 '\\-'（在岗显示为短横）。连同 number_format 一起复制，
    追加列才能保持同样的显示。表尾的 'Note:' / 'Prepared by' 等不会匹配真实 GID，
    留在 map 里也无害。
    """
    gid_col = _find_header_col(ws, ("GID",))
    start_col = _find_header_col(ws, (PR_START_HEADER,), contains=True)
    last_col = _find_header_col(ws, (PR_LAST_HEADER,), contains=True)
    missing = [
        name for name, col in (
            ("GID", gid_col), (PR_START_HEADER, start_col), (PR_LAST_HEADER, last_col),
        ) if col is None
    ]
    if missing:
        raise ExcelError(f"Payroll 报告未找到必需列：{missing}")

    mp: dict[str, tuple] = {}
    for r in range(1, ws.max_row + 1):
        raw = ws.cell(r, gid_col).value
        if raw is None:
            continue
        gid = str(raw).strip()
        if gid == "" or gid == "GID":
            continue
        s_cell = ws.cell(r, start_col)
        l_cell = ws.cell(r, last_col)
        mp[gid] = (s_cell.value, s_cell.number_format, l_cell.value, l_cell.number_format)
    return mp


def _clone_style(src, dst) -> None:
    """把 src 单元格的视觉样式复制到 dst（不含 number_format，单独设置）。"""
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)


def _insert_blank_cols(ws, at: int, n: int) -> None:
    """在第 ``at`` 列前插入 n 个空列，并把受影响的合并区右移 n。

    openpyxl 的 ``insert_cols`` 会搬动单元格的值与样式，但**不会**搬动合并区
    （留着会和新列重叠、令表头错位）。列宽则按列号保持不动——这与 Excel 手动
    插入列的表现一致（薪资块各列宽度留在原列号上），故无需处理。
    """
    old = [(m.min_col, m.min_row, m.max_col, m.max_row) for m in ws.merged_cells.ranges]
    # 先在插入前拆掉合并（此时各单元格都还在，unmerge 不会因新空列缺单元格而 KeyError）
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))
    ws.insert_cols(at, n)
    for c1, r1, c2, r2 in old:
        ws.merge_cells(
            start_row=r1, end_row=r2,
            start_column=c1 + n if c1 >= at else c1,
            end_column=c2 + n if c2 >= at else c2,
        )


def _process_gl_like(
    input_path: Path,
    output_path: Path,
    code_headers: tuple[str, ...],
    password: str,
) -> tuple[int, int]:
    """处理 GL / 13th 报告。返回 (匹配行数, 总数据行数)。"""
    wb = open_xlsx(input_path, password=password)
    ws = wb.active

    code_col = _find_col(ws, code_headers)
    if code_col is None:
        raise ExcelError(f"未找到 GL 列：尝试过 {code_headers}")
    ba_col = _find_col(ws, (BA_HEADER,))
    cc_col = _find_col(ws, (CC_HEADER,))
    pc_col = _find_col(ws, (PC_HEADER,))
    missing = [
        name for name, col in (
            (BA_HEADER, ba_col), (CC_HEADER, cc_col), (PC_HEADER, pc_col),
        ) if col is None
    ]
    if missing:
        raise ExcelError(f"未找到必需列：{missing}")

    touched = 0
    total = 0
    for r in range(2, ws.max_row + 1):
        gl = ws.cell(r, code_col).value
        if gl is None or str(gl).strip() == "":
            continue
        total += 1
        if _starts_with(gl, "2"):
            ws.cell(r, ba_col).value = BA_VALUE
            ws.cell(r, pc_col).value = PC_VALUE
            ws.cell(r, cc_col).value = None
            touched += 1

    _save_encrypted(wb, output_path, password)
    return touched, total


class VaVnReportTask(TaskBase):
    task_id = "va-vn-payroll-report"
    code    = "VA-VN-PAYROLL-REPORT"
    name    = "瓦里安越南 Payroll 报告加工"
    desc    = "GL CODE 2 填 BusinessArea/ProfitCenter · GL CODE 6 填 CostCenter · Variance 加入职/离职日"
    inputs  = ["xlsx"]

    def run(self, *, input_path: Path, output_dir: Path, options: dict) -> Iterator[TaskEvent]:
        opts = options or {}
        station = str(opts.get("station") or "").strip()
        password = str(opts.get("password") or PASSWORD)

        if station.upper() == "GL":
            yield from self._run_gl_or_13th(
                input_path, output_dir, password,
                kind="GL", code_headers=("G/L Acc", "GL Acc"),
            )
        elif station == "13th":
            yield from self._run_gl_or_13th(
                input_path, output_dir, password,
                kind="13th", code_headers=("G/L Code", "GL Code"),
            )
        elif station.lower() == "variance":
            yield from self._run_variance(input_path, output_dir, opts, password)
        else:
            # 没传 station：保留 placeholder 行为以兼容旧调用
            yield from self._run_legacy_placeholder(output_dir, opts)

    # ── GL / 13th ──────────────────────────────────────────────────────────

    def _run_gl_or_13th(
        self,
        input_path: Path,
        output_dir: Path,
        password: str,
        *,
        kind: str,
        code_headers: tuple[str, ...],
    ) -> Iterator[TaskEvent]:
        yield LogEvent(f"读取 {kind} 报告：{input_path.name}")
        yield ProgressEvent(done=0, total=2, note=f"{kind} · 解密读取")

        out_path = output_dir / input_path.name
        try:
            touched, total = _process_gl_like(input_path, out_path, code_headers, password)
        except ExcelError:
            raise  # 输入校验错误：消息已友好，交给 server 统一展示（不加 {kind} 前缀）
        except Exception as e:  # noqa: BLE001
            yield LogEvent(f"{kind} 处理失败：{e}", lvl="err")
            raise

        yield ProgressEvent(done=1, total=2, note=f"{kind} · 写出")
        yield LogEvent(
            f"{kind}：共 {total} 行，命中 GL CODE 2 开头 {touched} 行",
            lvl="ok",
        )
        yield LogEvent(f"写出：{out_path.name}", lvl="ok")
        yield str(out_path)
        yield ProgressEvent(done=2, total=2, note="done")

    # ── Variance ─────────────────────────────────────────────────────────

    def _run_variance(
        self, input_path: Path, output_dir: Path, options: dict, password: str
    ) -> Iterator[TaskEvent]:
        """Payroll(input_path) × Variance(options['input2']) 按 GID 关联。

        把 Payroll 的 Start date / Last working date 插入到 Variance 的
        Full Name 之后（第 5、6 列）。
        """
        period = options.get("period") or time.strftime("%b %Y")
        dashed = str(period).replace(" ", "-")

        # 服务器只校验主输入(input)存在；第二个文件得自己兜底。
        variance_in = options.get("input2")
        if not variance_in:
            raise ExcelError("缺少 Variance 报告（未收到第二个输入文件）")
        variance_path = Path(str(variance_in))
        if not variance_path.exists():
            raise ExcelError(f"Variance 报告不存在：{variance_path}")

        yield LogEvent(f"读取 Payroll：{input_path.name}")
        yield ProgressEvent(done=0, total=3, note="解密 Payroll")
        pw = open_xlsx(input_path, password=password, data_only=True)
        pw_ws = _find_payroll_sheet(pw)
        if pw_ws is None:
            raise ExcelError("Payroll 报告未找到含 GID / Start date 的工作表")
        payroll_map = _build_payroll_map(pw_ws)
        yield LogEvent(f"Payroll：建立 GID 索引 {len(payroll_map)} 条", lvl="ok")

        yield ProgressEvent(done=1, total=3, note="解密 Variance")
        yield LogEvent(f"读取 Variance：{variance_path.name}")
        vw = open_xlsx(variance_path, password=password)
        vw_ws = vw.active
        gid_col = _find_header_col(vw_ws, ("GID",))
        if gid_col is None:
            raise ExcelError("Variance 报告未找到 GID 列")
        name_col = _find_header_col(vw_ws, ("Full Name",))
        if name_col is None:
            raise ExcelError("Variance 报告未找到 Full Name 列")

        # 紧跟 Full Name 插入两列（gid_col 在其左侧，列号不受影响）
        start_out = name_col + 1
        last_out = start_out + 1
        _insert_blank_cols(vw_ws, start_out, 2)
        for col, title, width in (
            (start_out, OUT_START_HEADER, 22),
            (last_out, OUT_LAST_HEADER, 25),
        ):
            hdr = vw_ws.cell(1, col)
            hdr.value = title
            # 身份列(No./GID/…)的表头是纵向合并 1:2，两列日期照此对齐
            vw_ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
            _clone_style(vw_ws.cell(1, gid_col), hdr)
            vw_ws.column_dimensions[get_column_letter(col)].width = width

        total = 0
        matched = 0
        for r in range(1, vw_ws.max_row + 1):
            raw = vw_ws.cell(r, gid_col).value
            if raw is None:
                continue
            gid = str(raw).strip()
            if gid == "" or gid == "GID":
                continue
            total += 1
            rec = payroll_map.get(gid)
            if rec is None:
                continue
            s_val, s_fmt, l_val, l_fmt = rec
            # Payroll 用 0（配 '-' 格式）表示在岗 → 离职日留空
            if l_val in (None, "", 0, "0"):
                l_val, l_fmt = None, "General"
            src = vw_ws.cell(r, gid_col)
            for col, val, fmt in ((start_out, s_val, s_fmt), (last_out, l_val, l_fmt)):
                cell = vw_ws.cell(r, col)
                cell.value = val
                _clone_style(src, cell)
                cell.number_format = fmt
            matched += 1

        yield ProgressEvent(done=2, total=3, note="加密写出")
        out = output_dir / f"VarianVN_Variance_{dashed}_CDP.xlsx"
        _save_encrypted(vw, out, password)

        if matched == total:
            yield LogEvent(f"Variance：{total} 行全部按 GID 匹配", lvl="ok")
        else:
            yield LogEvent(
                f"Variance：{total} 行，匹配 {matched} 行，"
                f"{total - matched} 行未在 Payroll 找到 GID",
                lvl="warn",
            )
        yield LogEvent(f"写出：{out.name}", lvl="ok")
        yield str(out)
        yield ProgressEvent(done=3, total=3, note="done")

    # ── 旧 placeholder（无 station 时） ───────────────────────────────────

    def _run_legacy_placeholder(
        self, output_dir: Path, options: dict
    ) -> Iterator[TaskEvent]:
        period = options.get("period") or time.strftime("%b %Y")
        dashed = str(period).replace(" ", "-")
        subs = [("GL", 184 * 1024), ("13th", 96 * 1024), ("Variance", 142 * 1024)]
        total = len(subs)
        for i, (kind, bytes_) in enumerate(subs):
            time.sleep(0.1)
            yield ProgressEvent(done=i, total=total, note=kind)
            out = output_dir / f"VarianVN_{kind}_{dashed}_CDP.xlsx"
            with out.open("wb") as f:
                f.write(os.urandom(bytes_))
            yield LogEvent(f"写出：{out.name}", lvl="ok")
            yield str(out)
        yield ProgressEvent(done=total, total=total, note="done")
