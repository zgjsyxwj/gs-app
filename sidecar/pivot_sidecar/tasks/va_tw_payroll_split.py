"""瓦里安TW Payroll 账单拆分 (VA-TW-PAYROLL-SPLIT)

按预定义的 sheet 映射，把一个供应商 xlsx 拆成 4 个独立 xlsx：
  · 部門薪資總表 + 員工薪資表 → Varian_Salary Report.xlsx
  · 員工加班費明細表           → Varian_OT Details Report.xlsx
  · 保險資料明細               → Varian_Social Details Report.xlsx
  · 薪資差異分析表             → Varian_Variance Report.xlsx

文件名前缀 = options['period']，输出文件用 options['password'] 加密 (ECMA-376 Agile)。

公式策略 · 见 docs 里的 option B 决策：
- Salary / OT / Social 的公式都是 sheet 内部引用，直接 copy 即可。
- 薪資差異分析表 有 304 条 XLOOKUP 指向 員工資料匯出 (该 sheet 不在
  任何输出里 → 拆分后会变成 #REF!)。所以 Variance 输出时把这张
  sheet 的公式逐 cell 替换为 Excel 缓存的计算结果，相当于一个
  「时点快照」 — 不再能重算，但也不带走员工原始资料表。
"""
from __future__ import annotations
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

from .base import TaskBase, TaskEvent
from ..ipc import LogEvent, ProgressEvent
from ..xlsx import ExcelError, open_xlsx


# (输出文件名后缀, 保留的 sheet 列表) — 顺序与前端 SPLITS 一致
SPLITS: list[tuple[str, list[str]]] = [
    ("Varian_Salary Report.xlsx", ["部門薪資總表", "員工薪資表"]),
    ("Varian_OT Details Report.xlsx", ["員工加班費明細表"]),
    ("Varian_Social Details Report.xlsx", ["保險資料明細"]),
    ("Varian_Variance Report.xlsx", ["薪資差異分析表"]),
]

# 输出时需要把公式攤平为 cached values 的 sheet (见 docstring · option B)
FLATTEN_SHEETS: set[str] = {"薪資差異分析表"}


def _encrypt_in_place(path: Path, password: str) -> None:
    """对一个明文 xlsx 加密 (ECMA-376 Agile)，覆盖原文件。"""
    import msoffcrypto

    tmp = path.with_suffix(path.suffix + ".tmp")
    try:
        with open(path, "rb") as fin, open(tmp, "wb") as fout:
            msoffcrypto.OfficeFile(fin).encrypt(password, fout)
        tmp.replace(path)
    except Exception:
        if tmp.exists():
            tmp.unlink()
        raise


class VaTwPayrollSplitTask(TaskBase):
    task_id = "va-tw-payroll-split"
    code = "VA-TW-PAYROLL-SPLIT"
    name = "瓦里安TW Payroll 账单拆分"
    desc = "按 sheet 映射拆成 Salary/OT/Social/Variance 4 个独立工作簿 · 加密"
    inputs = ["xlsx"]

    def run(
        self, *, input_path: Path, output_dir: Path, options: dict
    ) -> Iterator[TaskEvent]:
        period = str(options.get("period") or "").strip()
        password = str(options.get("password") or "").strip()
        if not period:
            yield LogEvent("缺少 period (yyyyMM) 参数", lvl="err")
            return

        yield LogEvent(f"读取输入：{input_path.name}")

        # 先校验文件可打开 + 所有需要的 sheet 都在（坏文件在此抛 ExcelError）
        all_needed = [name for _, names in SPLITS for name in names]
        probe = open_xlsx(input_path, read_only=True)
        existing = set(probe.sheetnames)
        probe.close()
        missing = [n for n in all_needed if n not in existing]
        if missing:
            raise ExcelError(f"账单缺少预期 sheet：{', '.join(missing)}")

        total = len(SPLITS)
        for i, (out_suffix, keep) in enumerate(SPLITS):
            out_name = f"{period}_{out_suffix}"
            out_path = output_dir / out_name
            yield LogEvent(f"生成：{out_name}")

            # 每个输出都重新载入源文件 → 删除不需要的 sheet → 另存。
            # 这样比手工 copy 行更可靠，能完整保留样式 / 合并单元格 /
            # 列宽 / 行高 / 公式。源文件不会被修改。
            wb = load_workbook(input_path)
            keep_set = set(keep)
            for name in list(wb.sheetnames):
                if name not in keep_set:
                    del wb[name]
            # 保持 SPLITS 中声明的顺序
            order = {name: idx for idx, name in enumerate(keep)}
            wb._sheets.sort(key=lambda ws: order.get(ws.title, len(order)))

            # option B · 把指定 sheet 的公式攤平为 Excel 缓存值。前提是
            # 源檔上次由 Excel 写出 (会带 cached values)。如果是
            # LibreOffice / 程序生成的 xlsx 可能没有缓存 → 该 cell 落 None
            # 并 emit warn。
            flatten_here = [s for s in keep if s in FLATTEN_SHEETS]
            flattened_total = 0
            missing_cache = 0
            if flatten_here:
                # Not read_only: we need .coordinate on cells, and read_only
                # yields EmptyCell objects for the bounding-box padding which
                # don't expose it. Full load is fine for this file class.
                wb_v = load_workbook(input_path, data_only=True)
                try:
                    for sheet_name in flatten_here:
                        ws = wb[sheet_name]
                        ws_v = wb_v[sheet_name]
                        # Keep ALL cells incl. those with cached "" — XLOOKUP
                        # against an empty target column legitimately returns
                        # the empty string, and that's real data we mustn't
                        # silently drop to None.
                        cache = {
                            c.coordinate: c.value
                            for row in ws_v.iter_rows()
                            for c in row
                        }
                        for row in ws.iter_rows():
                            for cell in row:
                                v = cell.value
                                if isinstance(v, str) and v.startswith("="):
                                    cached = cache.get(cell.coordinate)
                                    if cached is None:
                                        missing_cache += 1
                                    cell.value = cached
                                    flattened_total += 1
                finally:
                    wb_v.close()
                if flattened_total:
                    if missing_cache:
                        yield LogEvent(
                            f"  攤平 {flattened_total} 个公式 → cached values · "
                            f"其中 {missing_cache} 个源檔未缓存计算结果 (已留空) · "
                            "建议在 Excel 内按 Ctrl+Alt+F9 强制重算后再保存源檔",
                            lvl="warn",
                        )
                    else:
                        yield LogEvent(
                            f"  攤平 {flattened_total} 个公式 → cached values",
                            lvl="info",
                        )

            wb.save(out_path)
            wb.close()

            if password:
                _encrypt_in_place(out_path, password)

            yield ProgressEvent(done=i + 1, total=total, note=out_name)
            yield str(out_path)

        yield LogEvent("拆分完成", lvl="ok")
